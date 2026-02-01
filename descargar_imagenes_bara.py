#!/usr/bin/env python3
# -*- coding: utf-8 -*-

import os
import re
import csv
import unicodedata
import zipfile
from urllib.parse import urlparse, parse_qs

import requests
from openpyxl import load_workbook

# ==========================
# CONFIGURACIÓN
# ==========================

EXCEL_PATH = r"C:\bara\LISTA DE PRECIOS NOVIEMBRE 13 2025 TRABAJO JC.xlsx"
OUTPUT_DIR = r"C:\bara\imagenes_bara"
LOG_CSV = os.path.join(OUTPUT_DIR, "log_descargas_imagenes.csv")
ZIP_PATH = os.path.join(OUTPUT_DIR, "imagenes_bara_importaciones.zip")

REQUEST_TIMEOUT = 20
MAX_RETRIES = 3

def ensure_dir(path: str):
    if not os.path.exists(path):
        os.makedirs(path, exist_ok=True)

def normalize_text(text: str, max_len: int = 80) -> str:
    if not text:
        return ""
    text = str(text).strip().upper()
    text = unicodedata.normalize("NFKD", text)
    text = text.encode("ascii", "ignore").decode("ascii")
    text = text.replace("/", "-")
    text = re.sub(r"\s+", "-", text)
    text = re.sub(r"[^A-Z0-9\-_]", "", text)
    return text[:max_len]

def extract_drive_id(url: str):
    if not url:
        return None
    m = re.search(r"/d/([a-zA-Z0-9_\-]+)/", url)
    if m:
        return m.group(1)
    parsed = urlparse(url)
    qs = parse_qs(parsed.query)
    if "id" in qs:
        return qs["id"][0]
    return None

def build_direct_url(url: str):
    fid = extract_drive_id(url)
    if fid:
        return f"https://drive.google.com/uc?export=download&id={fid}"
    return None

def guess_ext(ct: str):
    if not ct:
        return ".jpg"
    ct = ct.lower()
    if "jpeg" in ct:
        return ".jpg"
    if "png" in ct:
        return ".png"
    if "gif" in ct:
        return ".gif"
    if "webp" in ct:
        return ".webp"
    return ".jpg"

def unique_path(folder, basename, ext):
    candidate = os.path.join(folder, basename + ext)
    if not os.path.exists(candidate):
        return candidate
    i = 1
    while True:
        c = os.path.join(folder, f"{basename}_{i}{ext}")
        if not os.path.exists(c):
            return c
        i += 1

def download(url, dest):
    for attempt in range(1, MAX_RETRIES + 1):
        try:
            print(f"  → Intento {attempt}: {url}")
            with requests.get(url, stream=True, timeout=REQUEST_TIMEOUT) as r:
                if r.status_code != 200:
                    print(f"    ⚠️ HTTP {r.status_code}")
                    continue
                ct = r.headers.get("Content-Type", "")
                if dest.endswith(".tmp"):
                    ext = guess_ext(ct)
                    dest = dest[:-4] + ext
                with open(dest, "wb") as f:
                    for chunk in r.iter_content(8192):
                        if chunk:
                            f.write(chunk)
            print(f"    ✅ Guardado en: {dest}")
            return True, dest
        except Exception as e:
            print(f"    ⚠️ Error: {e}")
    return False, dest

def make_zip(folder, zip_path):
    print(f"\n📦 Empaquetando ZIP en {zip_path}")
    with zipfile.ZipFile(zip_path, "w", zipfile.ZIP_DEFLATED) as z:
        for root, _, files in os.walk(folder):
            for file in files:
                if file.lower().endswith((".jpg", ".jpeg", ".png", ".gif", ".webp")):
                    fp = os.path.join(root, file)
                    rp = os.path.relpath(fp, folder)
                    z.write(fp, rp)
    print("   ✔ ZIP listo.")

def main():
    print("📂 Cargando Excel:", EXCEL_PATH)
    ensure_dir(OUTPUT_DIR)

    wb = load_workbook(EXCEL_PATH, data_only=True)
    ws = wb.active
    print("📄 Hoja usada:", ws.title)

    # 1. Detectar encabezados
    header_row = None
    for row in ws.iter_rows(min_row=1, max_row=50):
        for c in row:
            val = str(c.value).upper().strip() if c.value else ""
            if "PRODUCTO / DESCRIPCION" in val or "PRODUCTO/DESCRIPCION" in val:
                header_row = c.row
                break
        if header_row:
            break

    if not header_row:
        print("❌ No encontré encabezados.")
        return

    col_desc = col_img = col_code = None
    for c in ws[header_row]:
        txt = str(c.value).strip().upper() if c.value else ""
        if "PRODUCTO" in txt and "DESCRIPCION" in txt:
            col_desc = c.column
        elif "VER IMAGEN" in txt:
            col_img = c.column
        elif txt == "CODIGO" or txt == "CÓDIGO":
            col_code = c.column

    if not (col_desc and col_img and col_code):
        print("❌ No encontré todas las columnas necesarias.")
        print(col_desc, col_img, col_code)
        return

    print(f"   ✔ DESCRIPCION col {col_desc}")
    print(f"   ✔ VER IMAGEN  col {col_img}")
    print(f"   ✔ CODIGO      col {col_code}")

    rows_log = []
    print("\n🚀 Descargando imágenes…\n")

    for row in ws.iter_rows(min_row=header_row + 1):
        desc = row[col_desc - 1].value
        codigo = row[col_code - 1].value
        img_cell = row[col_img - 1]

        desc = str(desc).strip() if desc else ""
        codigo = str(codigo).strip() if codigo else ""

        if not desc and not codigo:
            continue

        hyperlink = img_cell.hyperlink.target if img_cell.hyperlink else None

        if not hyperlink:
            print(f"⚠️ Sin link → {desc}")
            rows_log.append([desc, codigo, "", "", "SIN_LINK"])
            continue

        direct_url = build_direct_url(hyperlink)
        if not direct_url:
            print(f"⚠️ Sin ID Drive → {hyperlink}")
            rows_log.append([desc, codigo, hyperlink, "", "BAD_URL"])
            continue

        # Nombre final
        desc_n = normalize_text(desc)
        cod_n = normalize_text(codigo)

        base = cod_n if cod_n else "SIN_CODIGO"
        if desc_n:
            base = f"{cod_n}-{desc_n}"

        temp_path = unique_path(OUTPUT_DIR, base, ".tmp")

        ok, final_path = download(direct_url, temp_path)

        if ok:
            rows_log.append([desc, codigo, hyperlink, final_path, "OK"])
        else:
            rows_log.append([desc, codigo, hyperlink, final_path, "ERROR"])

    # Guardar LOG
    with open(LOG_CSV, "w", newline="", encoding="utf-8-sig") as f:
        w = csv.writer(f, delimiter=";")
        w.writerow(["DESCRIPCION", "CODIGO", "URL", "LOCAL", "STATUS"])
        w.writerows(rows_log)

    make_zip(OUTPUT_DIR, ZIP_PATH)

    print("\n🎉 FINALIZADO")
    print("Imágenes:", OUTPUT_DIR)
    print("Log:", LOG_CSV)
    print("ZIP:", ZIP_PATH)

if __name__ == "__main__":
    main()
